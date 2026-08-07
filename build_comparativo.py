# -*- coding: utf-8 -*-
"""
Gera comparativo_nv.js a partir de duas planilhas:
  1. "NV por Macro_2019 x 2024_1500g": <1500g por macro e UF
  2. "NV por Região_2019 x 2024": todos os NV por região de saúde, macro e UF
% de nascidos vivos ocorridos no território de residência, 2019 e 2024,
casando cada linha da planilha com o polígono correspondente do GEO.
"""
import glob as _glob
import json, os, re, sys, unicodedata
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\anapa\Downloads\NV por Macro_2019 x 2024_1500g (1).xlsx"
XLSX_RS = sys.argv[2] if len(sys.argv) > 2 else sorted(_glob.glob(r"C:\Users\anapa\Downloads\NV por Regi*o_2019 x 2024 22_03.xlsx"))[0]

UF_COD = {11:"RO",12:"AC",13:"AM",14:"RR",15:"PA",16:"AP",17:"TO",21:"MA",22:"PI",23:"CE",24:"RN",
          25:"PB",26:"PE",27:"AL",28:"SE",29:"BA",31:"MG",32:"ES",33:"RJ",35:"SP",41:"PR",42:"SC",
          43:"RS",50:"MS",51:"MT",52:"GO",53:"DF"}
UF_NOME = {"acre":"AC","alagoas":"AL","amazonas":"AM","amapa":"AP","bahia":"BA","ceara":"CE",
           "distrito federal":"DF","espirito santo":"ES","goias":"GO","maranhao":"MA","minas gerais":"MG",
           "mato grosso do sul":"MS","mato grosso":"MT","para":"PA","paraiba":"PB","pernambuco":"PE",
           "piaui":"PI","parana":"PR","rio de janeiro":"RJ","rio grande do norte":"RN","rondonia":"RO",
           "roraima":"RR","rio grande do sul":"RS","santa catarina":"SC","sergipe":"SE","sao paulo":"SP",
           "tocantins":"TO"}
ROMANOS = {"i":1,"ii":2,"iii":3,"iv":4,"v":5,"vi":6,"vii":7,"viii":8,"ix":9,"x":10,"xi":11,"xii":12,
           "xiii":13,"xiv":14,"xv":15,"xvi":16,"xvii":17,"xviii":18,"xix":19,"xx":20}

def norm(s):
    s = unicodedata.normalize("NFD", str(s or "")).lower()
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def numeros(s):
    """extrai números do nome (ordinais 1ª, romanos II, arábicos)"""
    n = set()
    for tk in norm(s).split():
        if tk.isdigit(): n.add(int(tk))
        elif tk in ROMANOS: n.add(ROMANOS[tk])
    m = re.findall(r"(\d+)\s*[ªº°a]", norm(s))
    for x in m: n.add(int(x))
    return n

PALAVRAS_VAZIAS = {"macro","macrorregiao","macrorregional","macrorregiao","de","da","do","saude","regiao","unica","uf",
                   "1","2","3","4","5","6","7","8","9"}
def tokens(s):
    return {t for t in norm(s).split() if t not in PALAVRAS_VAZIAS and t not in ROMANOS and len(t) > 2}

# ---------- GEO macros e regiões de saúde ----------
js = open(os.path.join(BASE, "dados_nv2026.js"), encoding="utf-8").read()
geo = json.loads(re.search(r"const GEO = (\{.*?\});", js, re.S).group(1))
def geo_nivel(nivel):
    out = []
    for t in geo["niveis"][nivel]:
        uf = UF_COD.get(int(str(t["id"])[:2]))
        out.append({"id": t["id"], "nome": t["nome"], "uf": uf})
    return out
geo_macros = geo_nivel("macro")
geo_rs = geo_nivel("rs")

# ---------- planilha ----------
wb = openpyxl.load_workbook(XLSX, data_only=True)
def parse_aba(nome):
    ws = wb[nome]
    macros, ufs, pendentes = [], {}, []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rotulo = str(row[0] or "").strip()
        if not rotulo: continue
        try:
            oc, res, pct = float(row[1]), float(row[2]), float(row[3])
        except (TypeError, ValueError):
            continue
        if rotulo.upper().startswith("UF"):
            sigla = UF_NOME.get(norm(re.sub(r"^UF\s*-\s*", "", rotulo, flags=re.I)))
            if sigla:
                ufs[sigla] = {"oc": oc, "res": res, "pct": pct}
                for p in pendentes: p["uf"] = sigla
                macros.extend(pendentes)
                pendentes = []
        elif "TOTAL" in rotulo.upper() or "BRASIL" in rotulo.upper():
            continue
        else:
            pendentes.append({"rotulo": rotulo, "oc": oc, "res": res, "pct": pct, "uf": None})
    macros.extend(p for p in pendentes if p["uf"])
    return macros, ufs

# ---------- casamento planilha -> polígono ----------
GENERICAS = {"macro","macrorregiao","macrorregional","regiao","de","da","do","saude","uf"}
def chave(s, uf=None):
    """nome normalizado sem palavras genéricas e sem o sufixo da UF"""
    n = norm(re.sub(r"-\s*[A-Z]{2}\s*$", "", str(s)))
    if uf: n = re.sub(rf"\b{uf.lower()}\b", " ", n)
    return " ".join(t for t in n.split() if t not in GENERICAS)

def casa(macros_plan, universo=None):
    resultado, sem_par = {}, []
    por_uf = {}
    for g in (universo if universo is not None else geo_macros): por_uf.setdefault(g["uf"], []).append(g)
    usados, feito = set(), set()

    def atribui(mp, g):
        usados.add(g["id"]); feito.add(id(mp))
        resultado[g["id"]] = {"oc": mp["oc"], "res": mp["res"], "pct": mp["pct"], "plan": mp["rotulo"]}

    # fase 0: UF com macro única  |  fase 1: nome normalizado idêntico
    for mp in macros_plan:
        cands = por_uf.get(mp["uf"], [])
        if len(cands) == 1:
            atribui(mp, cands[0])
    for mp in macros_plan:
        if id(mp) in feito: continue
        k = chave(mp["rotulo"], mp["uf"])
        exatos = [g for g in por_uf.get(mp["uf"], []) if g["id"] not in usados and chave(g["nome"], mp["uf"]) == k]
        if len(exatos) == 1:
            atribui(mp, exatos[0])

    # fase 2: número (romano/ordinal) e tokens em comum
    for mp in macros_plan:
        if id(mp) in feito: continue
        cands = [g for g in por_uf.get(mp["uf"], []) if g["id"] not in usados]
        alvo = None
        nplan = numeros(mp["rotulo"]); tplan = tokens(mp["rotulo"])
        com_num = [g for g in cands if nplan and numeros(g["nome"]) & nplan]
        if len(com_num) == 1: alvo = com_num[0]
        else:
            com_tok = [(len(tokens(g["nome"]) & tplan), g) for g in cands]
            com_tok = [x for x in com_tok if x[0] > 0]
            if com_tok:
                com_tok.sort(key=lambda x: -x[0]); alvo = com_tok[0][1]
            elif len(com_num) > 1:
                alvo = com_num[0]
            elif len(cands) == 1:
                alvo = cands[0]
        if alvo: atribui(mp, alvo)
        else: sem_par.append((mp["uf"], mp["rotulo"]))
    return resultado, sem_par

# ---------- planilha por região de saúde (todos os NV: UF -> macro -> RS) ----------
wb_rs = openpyxl.load_workbook(XLSX_RS, data_only=True)
def parse_aba_rs(nome):
    ws = wb_rs[nome]
    rss, macros, ufs = [], [], {}
    uf_atual = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        rotulo = str(row[0] or "").strip()
        if not rotulo: continue
        try:
            oc, res, pct = float(row[1]), float(row[2]), float(row[3])
        except (TypeError, ValueError):
            continue
        up = rotulo.upper()
        if up.startswith("UF"):
            uf_atual = UF_NOME.get(norm(re.sub(r"^UF\s*-\s*", "", rotulo, flags=re.I)))
            if uf_atual: ufs[uf_atual] = {"oc": oc, "res": res, "pct": pct}
        elif "TOTAL" in up or up.startswith("BRASIL"):
            continue
        elif up.startswith("MACRORREGI"):
            nome_m = re.sub(r"^Macrorregi\w+\s+de\s+Sa\w+\s*-\s*", "", rotulo, flags=re.I)
            macros.append({"rotulo": nome_m, "oc": oc, "res": res, "pct": pct, "uf": uf_atual})
        else:
            rss.append({"rotulo": rotulo, "oc": oc, "res": res, "pct": pct, "uf": uf_atual})
    return rss, macros, ufs

saida = {"anos": ["2019", "2024"],
         "m1500": {"macro": {}, "uf": {}},
         "todos": {"rs": {}, "macro": {}, "uf": {}}}

for ano in ("2019", "2024"):
    # <1500g por macro
    macros_plan, ufs_plan = parse_aba(ano)
    casados, sem_par = casa(macros_plan)
    print(f"[<1500g {ano}] macros: {len(macros_plan)} | casadas: {len(casados)} | sem par: {len(sem_par)}")
    for uf, r in sem_par: print(f"    SEM PAR: {uf} | {r}")
    for gid, v in casados.items():
        saida["m1500"]["macro"].setdefault(str(gid), {})[ano] = [v["oc"], v["res"], v["pct"]]
    for uf, v in ufs_plan.items():
        saida["m1500"]["uf"].setdefault(uf, {})[ano] = [v["oc"], v["res"], v["pct"]]

    # todos os NV por RS / macro / UF
    rss_plan, macros_plan2, ufs_plan2 = parse_aba_rs(ano)
    cas_rs, sem_rs = casa(rss_plan, geo_rs)
    cas_m2, sem_m2 = casa(macros_plan2, geo_macros)
    print(f"[todos {ano}] RS: {len(rss_plan)} | casadas: {len(cas_rs)} | sem par: {len(sem_rs)}"
          f" || macros: {len(macros_plan2)} | casadas: {len(cas_m2)} | sem par: {len(sem_m2)}")
    for uf, r in (sem_rs + sem_m2): print(f"    SEM PAR: {uf} | {r}")
    for gid, v in cas_rs.items():
        saida["todos"]["rs"].setdefault(str(gid), {})[ano] = [v["oc"], v["res"], v["pct"]]
    for gid, v in cas_m2.items():
        saida["todos"]["macro"].setdefault(str(gid), {})[ano] = [v["oc"], v["res"], v["pct"]]
    for uf, v in ufs_plan2.items():
        saida["todos"]["uf"].setdefault(uf, {})[ano] = [v["oc"], v["res"], v["pct"]]

with open(os.path.join(BASE, "comparativo_nv.js"), "w", encoding="utf-8") as fh:
    fh.write("/* Gerado por build_comparativo.py: NV ocorridos no território de residência (todos e <1500g) */\nconst COMP = ")
    json.dump(saida, fh, ensure_ascii=False, separators=(",", ":"))
    fh.write(";\n")
print("gravado comparativo_nv.js")
