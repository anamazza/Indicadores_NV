# -*- coding: utf-8 -*-
"""Le as planilhas de Indicadores_Assistenciais/ (perfil sociodemografico SINASC,
internacoes obstetricas/neonatais SIH, morbidade neonatal) e devolve um dicionario
por CNES para o build_dados.py anexar as maternidades do painel.

Formato de cada bloco (chave = CNES com 7 digitos):
    {"origem": "APOIADAS", "nome": "<cabecalho da planilha>",
     "secoes": [{"titulo": "PERFIL ...", "linhas":
                 [{"rotulo": str, "fmt": "pct|int|dias", "valores": [7 valores 2019-2025]}]}]}

Regras de leitura (estrutura verificada em 26/08/2026):
- blocos lado a lado nas abas por regiao; cabecalho "NOME - CNES: NNNNNNN" na linha 1;
- linha 2 = "Indicador", 2019..2025; secoes = rotulo TODO em maiusculas sem valores;
- valores ja vem numericos (percentuais na escala 0-100, tempos em dias).
Os arquivos "_reparada" sao os oficiais com o drawing fantasma removido (LEIA-ME de la).
"""
import os
import re

from openpyxl import load_workbook

ARQUIVOS = [
    ("APOIADAS", "Indicadores_APOIADAS_reparada.xlsx"),
    ("EBSERH", "Indicadores_EBSERH_reparada.xlsx"),
    ("QUALINEO", "Indicadores_QUALINEO_reparada.xlsx"),
]
ANOS = list(range(2019, 2026))


def _fmt(rotulo):
    r = rotulo.lower()
    if r.startswith(("distribuição percentual", "porcentagem", "taxa de ocupação")):
        return "pct"
    if r.startswith("tempo médio"):
        return "dias"
    return "int"


def carregar(pasta):
    """Devolve (blocos_por_cnes, avisos)."""
    blocos, avisos = {}, []
    for origem, arq in ARQUIVOS:
        wb = load_workbook(os.path.join(pasta, arq), data_only=True)
        for ws in wb.worksheets:
            linhas = list(ws.iter_rows(values_only=True))
            if len(linhas) < 3:
                continue
            r1 = linhas[0]
            cols = [i for i, v in enumerate(r1)
                    if isinstance(v, str) and re.search(r"CNES:?\s*\d{7}", v)]
            for c0 in cols:
                cab = str(r1[c0]).strip()
                cnes = re.search(r"CNES:?\s*(\d{7})", cab).group(1)
                anos = [v for v in linhas[1][c0 + 1:c0 + 8]]
                if [int(a) for a in anos if a is not None] != ANOS:
                    avisos.append(f"{arq}/{ws.title} {cnes}: anos inesperados {anos}")
                    continue
                if cnes in blocos:
                    avisos.append(f"{cnes} repetido ({origem} x {blocos[cnes]['origem']}) — mantido o primeiro")
                    continue
                secoes = []
                for ln in linhas[2:]:
                    rot = ln[c0]
                    if rot is None or not str(rot).strip():
                        continue
                    rot = str(rot).strip()
                    vals = list(ln[c0 + 1:c0 + 8])
                    if rot == rot.upper() and all(v is None for v in vals):
                        secoes.append({"titulo": rot, "linhas": []})
                        continue
                    if not secoes:
                        avisos.append(f"{arq}/{ws.title} {cnes}: linha fora de secao: {rot[:40]}")
                        continue
                    secoes[-1]["linhas"].append(
                        {"rotulo": rot, "fmt": _fmt(rot), "valores": vals})
                nome = re.sub(r"\s*-\s*CNES.*$", "", cab).strip()
                blocos[cnes] = {"origem": origem, "nome": nome, "secoes": secoes}
    return blocos, avisos


if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding="utf-8")
    base = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                        "Indicadores_Assistenciais")
    bl, av = carregar(base)
    print(len(bl), "blocos")
    seccount = {}
    for b in bl.values():
        for s in b["secoes"]:
            seccount[s["titulo"]] = seccount.get(s["titulo"], 0) + 1
    for t, n in seccount.items():
        print(f"  {n:3d}x {t}")
    for a in av:
        print("AVISO:", a)
